import os
from openpyxl import Workbook, load_workbook


# menuList = ["Create File", "Write File", "Read File", "Rename File", "Delete File", "Exit"]
menuList = ["Create File", "Write File", "Read File", "Exit"]


def menu():
    print("Main Menu:")
    for i in range(1, len(menuList)+1):
        print(f"{i}: {menuList[i-1]}")


def check_file_exists(filename):
    try:
        if os.path.exists(filename):
            return True
        else:
            raise FileNotFoundError("File Not Found")
    except FileNotFoundError as e:
        return e


def create_file(filename, header):
    try:
        if os.path.exists(filename):
            raise FileExistsError("File Already Exists")
        else:
            wb = Workbook()
            sheet = wb.active
            for i in range(len(header)):
                # print(header[i].strip())
                sheet.cell(row=1, column=i+1).value = header[i].strip()
            wb.save(filename)
            return "File Created Successfully"
    except FileExistsError as e:
        return e


def write_data(filename):
    wb = load_workbook(filename)
    sheet = wb.active

    values = []
    for i in range(1, sheet.max_column + 1):
        values.append(input(f"enter value for {sheet.cell(row=1, column=i).value}: "))

    row_number = sheet.max_row+1

    for i in range(1, sheet.max_column + 1):
        sheet.cell(row=row_number, column=i).value = values[i-1]

    choice = input("do you want to add another data? (y/n): ")

    wb.save(filename)

    if choice.lower() == 'y':
        write_data(filename)

    return "File Updated Successfully !!"


def read_data(filename):
    wb = load_workbook(filename)
    sheet = wb.active

    rows = sheet.max_row
    cols = sheet.max_column

    for rowNum in range(1, rows+1):
        for colNum in range(1, cols+1):
            print(f"{sheet.cell(row=rowNum, column=colNum).value}, ", end="")
        print()


def rename_file(filename, newname):
    try:
        if os.path.exists(filename):
            os.rename(filename, newname)
            return "File Renamed Successfully"
        else:
            raise FileNotFoundError("Old File Not Found")
    except FileNotFoundError as e:
        return e


# def delete_file(filename):
#     try:
#         if os.path.exists(filename):
#             confirm = input("Are you sure you want to delete? (y/n): ")
#             if confirm == "y" or confirm == "Y":
#                 os.remove(filename)
#                 return "File Deleted Successfully !!"
#             elif confirm == "n" or confirm == "N":
#                 return "File deletion cancelled"
#             else:
#                 raise IOError("Invalid Input")
#         else:
#             raise FileNotFoundError("File Not Found")
#     except Exception as e:
#         return e
