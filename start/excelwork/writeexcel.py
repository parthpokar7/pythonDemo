from openpyxl import Workbook, load_workbook


# # enter data in particular row
# filename = input("enter file name: ")
#
# wb = Workbook()
# sheet = wb.active
#
# rowNum = int(input("enter row number: "))
# data = input("enter value: ")
# if rowNum <= 0:
#     print("enter valid row number")
# else:
#     for i in range(5):
#         sheet.cell(row=rowNum, column=i+1).value = data
#
# wb.save(f"{filename}.xlsx")


# # enter data in particular column
# filename = input("enter file name: ")
#
# wb = Workbook()
# sheet = wb.active
#
# data = input("enter value: ")
#
# for i in range(5):
#     sheet.cell(row=i+1, column=5).value = data
#
# wb.save(f"{filename}.xlsx")


#read perticular column value
wb = load_workbook("column.xlsx")
sheet = wb.active

colname = input("enter column name: ")

rows = sheet.max_row
print(sheet[colname+"1"].value)
for cell in sheet[colname]:
    print(cell.value)

